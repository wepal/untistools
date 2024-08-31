import io
from pathlib import Path
import PIL
import openpyxl
import openpyxl.drawing

image_folder = Path(r"C:\Users\werne\OneDrive - Erich Fried Realgymnasium\2C\Fotos-2C")
outfilename = image_folder.parent/"Liste-2C.xlsx"
row_height = 15 #pt
column_width = 4 #chars
image_height = 20 #pixels

def to_openpyxl_img(img:PIL.Image):
    #https://foss.heptapod.net/openpyxl/openpyxl/-/issues/1477
    buf = io.BytesIO()
    img.save(buf, format='png')
    img = PIL.Image.open(buf)
    return openpyxl.drawing.image.Image(img)

workbook = openpyxl.Workbook()
sheet = workbook.active
for i, fname in enumerate(image_folder.glob('*.jpg')):
    img = PIL.Image.open(fname)
    img = img.crop((0, img.width//8, img.width, img.height-img.width//3))
    img = to_openpyxl_img(img)
    img.width,img.height = image_height*img.width/img.height,image_height
    sheet.add_image(img, f'B{i+1}')
    sheet.row_dimensions[i+1].height = row_height
    parts = fname.stem.split('_')
    sheet.cell(row=i+1, column=1).value = int(parts[0])
    sheet.cell(row=i+1, column=3).value = parts[1]
    sheet.cell(row=i+1, column=4).value = parts[2]
        
sheet.column_dimensions['A'].width = column_width
sheet.column_dimensions['B'].width = column_width
    
workbook.save(outfilename)